#!/usr/bin/python
#-* coding:UTF-8 -*

import pymysql
from pymysql.cursors import DictCursor, Cursor

import subprocess
from subprocess import Popen, PIPE, STDOUT, call

import json
import csv
import os
import codecs
import sys

import xlwt
import xlrd
from xlutils.copy import copy 

from openpyxl import Workbook  

reload(sys)
# os.system("ls")
sys.setdefaultencoding('utf-8')
abspath = os.path.abspath(sys.argv[0])
dname = os.path.dirname(abspath)
os.chdir(dname)


def open_excel(file='file.xls'):  
    try:  
        data = xlrd.open_workbook(file)  
        return data  
    except Exception,e:  
        print str(e)  

def excel_table_byindex(file='file.xls',colnameindex=0,by_index=0):  
    data = open_excel(file)  
    table = data.sheets()[by_index]  
    nrows = table.nrows #行数  
    ncols = table.ncols #列数  
    colnames = table.row_values(colnameindex) #某一行数据  
    list = []  
    for rownum in range(1,nrows):  
        row = table.row_values(rownum)#以列表格式输出  
        if row:  
            app = {}  
            for i in range(len(colnames)):  
                app[colnames[i]] = row[i]  
            list.append(app)#向列表中插入字典类型的数据  
    return list  

def writeToExcel():
    # 创建一个Workbook对象，这就相当于创建了一个Excel文件  
    book = xlwt.Workbook(encoding='utf-8', style_compression=0)  
    
    # 创建一个sheet对象，一个sheet对象对应Excel文件中的一张表格。  
    # 在电脑桌面右键新建一个Excel文件，其中就包含sheet1，sheet2，sheet3三张表  
    sheet = book.add_sheet('aa', cell_overwrite_ok=True)    # 其中的aa是这张表的名字  
    
    # 向表aa中添加数据  
    sheet.write(0, 0, 'EnglishName')    # 其中的'0, 0'指定表中的单元，'EnglishName'是向该单元写入的内容  
    sheet.write(1, 0, 'Marcovaldo')  
    txt1 = '中文名字'  
    sheet.write(0, 1, txt1.decode('utf-8')) # 此处需要将中文字符串解码成unicode码，否则会报错  
    txt2 = '马可瓦多'  
    sheet.write(1, 1, txt2.decode('utf-8'))  
  
    # 最后，将以上操作保存到指定的Excel文件中  
    book.save(r'e:\try1.xls') #在字符串前加r，声明为raw字符串，这样就不会处理其中的转义了。否则，可能会报错

def readFromExcel(file,colnameindex=0,by_index=0):
    data = open_excel(file)
    table = data.sheets()[by_index]  
    nrows = table.nrows #行数  
    ncols = table.ncols #列数
    colnames = table.col_values(colnameindex) #某一行数据  
    
    # for rownum in range(1,nrows):  
    #    row = table.row_values(rownum)#以列表格式输出  
    #     if row:

def replaceValue(file, colName="lang2", by_index=0):
    # for k in range(sList):
    #     v = sList[k]
    data = open_excel(file)
    wb = copy(data)  
    ws = wb.get_sheet(0)  
    ws.write(0, 3, 'EnglishName')    # 其中的'0, 0'指定表中的单元，'EnglishName'是向该单元写入的内容  
    ws.write(1, 3, 'Marcovaldo')  
    txt1 = '中文名字'  
    ws.write(2, 3, txt1.decode('utf-8')) # 此处需要将中文字符串解码成unicode码，否则会报错  
    txt2 = '马可瓦多'  
    ws.write(3, 3, txt2.decode('utf-8'))  
    wb.save("test.xls")  

def openpyxlTest():
    workbook = Workbook()  
    booksheet = workbook.active     #获取当前活跃的sheet,默认是第一个sheet  
    #存第一行单元格cell(1,1)  
    booksheet.cell(1,1).value = 6   #这个方法索引从1开始  
    booksheet.cell("B1").value = 7  
    #存一行数据  
    booksheet.append([11,87])  
    workbook.save("test_openpyxl.xlsx") 

def main():
    print("hello")

    # xlsfile = "fanyi2.xlsx"
    # # readFromExcel(xlsfile)
    # sList = excel_table_byindex(xlsfile)

    # file2 = "duoyuyan2.xlsx"
    # replaceValue(file2)

    openpyxlTest()

if __name__ == '__main__':
    main()