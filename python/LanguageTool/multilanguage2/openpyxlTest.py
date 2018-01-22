#!/usr/bin/python
#-* coding:UTF-8 -*

import json
import csv
import os
import codecs
import sys

from openpyxl import load_workbook

def readFromExcel():
    wb = load_workbook("duoyuyan2.xlsx")
    print(wb.sheetnames)    # ['Sheet1', 'Sheet2', 'Sheet3']
    sheet = wb.get_sheet_by_name("Sheet1")
    print(sheet["C3"].value)
    sheet["C3"] = "hello world"

    wb.save('newexcel.xlsx')
# def writeToExcel():

def main():
    print("hello openpyxl")
    readFromExcel()
    # writeToExcel()

if __name__ == '__main__':
    main()
