#!/usr/bin/python
#-* coding:UTF-8 -*

import json
import csv
import os
import codecs
import sys

reload(sys)
sys.setdefaultencoding('utf-8')
abspath = os.path.abspath(sys.argv[0])
dname = os.path.dirname(abspath)
os.chdir(dname)

from openpyxl import load_workbook

# ####################config#########################
# 翻译文件
translateFile = u"【火拼】1226-Thai Vietnamese Indonesian.xlsx"
# 多语言文件
languageFile = u"【火拼】多语言版本翻译表20180117.xlsx"
# 保存文件
saveFile = "replace.xlsx"
# 替换语言
languages = []
languages.append("lang2")
languages.append("lang3")
languages.append("lang4")
###################################################

# translateFile = "fanyi2.xlsx"
# languageFile = "duoyuyan2.xlsx"
# saveFile = "newexcel3.xlsx"

def open_excel(excelFile):  
    try:  
        wb = load_workbook(excelFile)
        return wb  
    except Exception,e:  
        print str(e)

def readFromExcel(excelFile):
    wb = open_excel(excelFile)
    if not wb:
        return
    sheet = wb["Sheet1"]
    nrow=sheet.max_row   #获取行数
    table = {}
    for i in range(2, nrow):
        row = sheet[i]
        print "readFromExcel: " + row[0].value
        table[row[0].value] = row

    return table    

def writeToExcel(excelFile, data, cols):
    wb = open_excel(excelFile)
    if not wb:
        return
    sheet = wb["Sheet1"]
    pos = []
    for i in range(len(sheet[1])):
        # print i
        row = sheet[1][i]
        # print value.value
        for v in cols:
            if v == row.value:
                pos.append(i)

    nrow=sheet.max_row   #获取行数
    for i in range(2, nrow ) :
        row = sheet[i]
        key = row[0].value
        # print "writeToExcel: " + key
        try: 
            # print data[key][pos].value
            dataRow = data[key]
            for p in pos:
                row[p].value = dataRow[p].value
        except Exception,e:  
            print "no key = " + str(e)
            for p in pos:
                row[p].value = 0

    wb.save(saveFile)
    print "=============== repace success! save to " + saveFile + " ================"

def main():
    # print("hello openpyxl")
    data = readFromExcel(translateFile)
    writeToExcel(languageFile, data, languages)

if __name__ == '__main__':
    main()
